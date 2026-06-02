"""
Amazon 货件发票导出模块

简介: 根据 Amazon 货件信息生成发票模板 xlsx，支持单货件导出与批量整理任务。

前端接口:
  - GET  /api/amazon/shipments/<id>/invoice/export        导出单个货件发票
  - POST /api/amazon/invoices/organize                    提交批量整理任务
  - GET  /api/amazon/invoices/organize/tasks              查询任务列表
  - GET  /api/amazon/invoices/organize/tasks/<id>         查询任务进度
  - GET  /api/amazon/invoices/organize/tasks/<id>/download 下载结果 zip
  - DELETE /api/amazon/invoices/organize/tasks/<id>       删除任务
  - POST /api/amazon/invoices/organize/tasks/<id>/cancel  取消任务
  - POST /api/amazon/invoices/organize/tasks/<id>/retry   重试任务

详细:
  1. 单货件导出：根据 shipment_id + shop_id + provider_id 查询货件箱规、产品信息、
     匹配对应货代的发票模板，填充数据后返回 xlsx。
  2. 批量整理：根据 inbound_plan_id 提交异步任务，后台 Worker 遍历该计划下所有货件，
     逐个生成 xlsx 并打包为 zip；前端轮询任务进度后下载。
  3. 后台 Worker 在 app 启动时（app.py）通过 _start_invoice_workers() 拉起，
     以守护线程运行，每 3 秒轮询 pending 任务。
"""
import io
import json
import os
import shutil
import threading
import time
import uuid
import urllib.request
from datetime import datetime, timedelta
from decimal import Decimal
from urllib.parse import urlparse

from flask import Blueprint, request, jsonify, send_file
from blueprints.user_auth import login_required, permission_required
from services.mysql_service import get_db_connection
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

amazon_invoice_export_bp = Blueprint('amazon_invoice_export', __name__, url_prefix='/api')

# ============================================================
# 常量
# ============================================================

TEMPLATES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    'static', 'shipping_invoice_templates'
)

INVOICE_TASK_DIR = os.path.join('static', 'shipping_invoice_templates', 'task')


# ============================================================
# 前端接口
# ============================================================

@amazon_invoice_export_bp.route('/amazon/shipments/<shipment_id>/invoice/export', methods=['GET'])
@login_required
@permission_required('amazon_invoice_export:export')
def export_shipment_invoice(shipment_id):
    """
    导出单个货件发票 xlsx

    简介: 根据货件编号生成对应货代的发票模板并返回 Excel 文件。

    详细:
      - 查询货件箱规及关联产品信息
      - 匹配货代发票模板配置
      - 填充数据并插入产品图片（本地路径或远程 URL）
      - 返回 .xlsx 文件流供前端下载

    查询参数:
        shop_id     (必填) 店铺ID
        provider_id (必填) 货代ID，决定使用哪套发票模板
    """
    try:
        shop_id = _require_shop_id()
        provider_id = request.args.get('provider_id', '').strip()
        if not provider_id:
            return jsonify({"status": "error", "message": "缺少必要参数: provider_id"}), 400
        try:
            provider_id = int(provider_id)
        except ValueError:
            return jsonify({"status": "error", "message": "provider_id 必须是整数"}), 400
        _, data = _fetch_shipment_invoice_data(shop_id=shop_id, shipment_id=shipment_id)
        if not data:
            return jsonify({"status": "error", "message": "该货件没有箱子数据"}), 404

        xlsx_io = _build_excel(data, provider_id=provider_id)
        filename = f"{shipment_id}.xlsx"

        return send_file(
            xlsx_io,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception as e:
        print(f"[Invoice Export] 导出异常: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize', methods=['POST'])
@login_required
@permission_required('amazon_invoice_export:export')
def submit_invoice_organize_task():
    """
    提交货件发票批量整理任务

    简介: 根据入库计划 ID 提交异步任务，后台自动生成该计划下所有货件的发票并打包为 zip。

    详细:
      - 先清理过期任务，避免堆积
      - 创建任务记录（状态 pending）
      - 后台 Worker 自动拾取并执行

    请求体 (JSON):
        shop_id         (必填) 店铺ID
        inbound_plan_id (必填) 入库计划ID
        provider_id     (必填) 货代ID

    返回:
        { status, message, data: { task_id } }
    """
    try:
        data = request.get_json() or {}
        shop_id = data.get('shop_id')
        inbound_plan_id = (data.get('inbound_plan_id') or '').strip()
        provider_id = data.get('provider_id')

        if not shop_id:
            return jsonify({"status": "error", "message": "shop_id 不能为空"}), 400
        try:
            shop_id = int(shop_id)
        except ValueError:
            return jsonify({"status": "error", "message": "shop_id 格式错误"}), 400

        if not inbound_plan_id:
            return jsonify({"status": "error", "message": "inbound_plan_id 不能为空"}), 400

        if not provider_id:
            return jsonify({"status": "error", "message": "provider_id 不能为空"}), 400
        try:
            provider_id = int(provider_id)
        except ValueError:
            return jsonify({"status": "error", "message": "provider_id 必须是整数"}), 400

        _cleanup_old_invoice_tasks()
        task_id = str(uuid.uuid4())
        _create_invoice_task(task_id, shop_id, inbound_plan_id, provider_id)

        return jsonify({
            "status": "success",
            "message": "任务已提交",
            "data": {"task_id": task_id}
        })

    except Exception as e:
        print(f"[Invoice Organize] 提交任务异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize/tasks', methods=['GET'])
@login_required
@permission_required('amazon_invoice_export:export')
def list_invoice_organize_tasks():
    """
    查询发票整理任务列表

    简介: 分页查询指定店铺下的任务，可按状态过滤，已完成任务附带下载链接。

    查询参数:
        shop_id   (必填) 店铺ID
        status    (可选) 按状态过滤: pending / running / completed / failed / cancelled
        page      (可选) 页码，默认 1
        page_size (可选) 每页条数，默认 10，最大 100

    返回:
        { status, data: { list, total, page, page_size } }
    """
    try:
        shop_id = request.args.get('shop_id', '').strip()
        if not shop_id:
            return jsonify({"status": "error", "message": "shop_id 不能为空"}), 400
        try:
            shop_id = int(shop_id)
        except ValueError:
            return jsonify({"status": "error", "message": "shop_id 格式错误"}), 400

        status = request.args.get('status', '').strip() or None
        page = max(1, int(request.args.get('page', 1)))
        page_size = max(1, min(100, int(request.args.get('page_size', 10))))

        result = _list_invoice_tasks(shop_id, status=status, page=page, page_size=page_size)

        for item in result['list']:
            item['created_at'] = item['created_at'].isoformat() if item['created_at'] else None
            item['completed_at'] = item['completed_at'].isoformat() if item['completed_at'] else None
            item['expired_at'] = item['expired_at'].isoformat() if item['expired_at'] else None
            if item['status'] == 'completed' and item['result_path']:
                item['download_url'] = f"/api/amazon/invoices/organize/tasks/{item['id']}/download"
            else:
                item['download_url'] = None

        return jsonify({"status": "success", "data": result})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize/tasks/<task_id>', methods=['GET'])
@login_required
@permission_required('amazon_invoice_export:export')
def get_invoice_organize_task(task_id):
    """
    查询发票整理任务进度

    简介: 根据任务 ID 获取状态、进度、完成时间等信息，前端可轮询此接口。

    返回:
        { status, data: { task_id, status, progress, message, created_at, completed_at, download_url } }
    """
    try:
        task = _get_invoice_task(task_id)
        if not task:
            return jsonify({"status": "error", "message": "任务不存在"}), 404

        result = {
            "task_id": task['id'],
            "status": task['status'],
            "progress": task['progress'],
            "message": task['message'] or '',
            "created_at": task['created_at'].isoformat() if task['created_at'] else None,
            "completed_at": task['completed_at'].isoformat() if task['completed_at'] else None,
        }

        if task['status'] == 'completed' and task['result_path']:
            result['download_url'] = f"/api/amazon/invoices/organize/tasks/{task_id}/download"
        else:
            result['download_url'] = None

        return jsonify({"status": "success", "data": result})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize/tasks/<task_id>/download', methods=['GET'])
@login_required
@permission_required('amazon_invoice_export:export')
def download_invoice_organize_task(task_id):
    """
    下载发票整理结果 zip

    简介: 任务完成后，下载包含所有货件 xlsx 的压缩包。

    详细:
      - 仅 completed 状态可下载
      - zip 内每个货件一个 xlsx 文件，文件名为 <shipment_id>.xlsx
    """
    try:
        task = _get_invoice_task(task_id)
        if not task:
            return jsonify({"status": "error", "message": "任务不存在"}), 404

        if task['status'] != 'completed':
            return jsonify({"status": "error", "message": "任务尚未完成"}), 400

        result_path = task['result_path']
        if not result_path or not os.path.exists(result_path):
            return jsonify({"status": "error", "message": "结果文件不存在"}), 404

        return send_file(
            result_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"{task['inbound_plan_id']}_invoices.zip"
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize/tasks/<task_id>', methods=['DELETE'])
@login_required
@permission_required('amazon_invoice_export:export')
def delete_invoice_organize_task(task_id):
    """
    删除发票整理任务

    简介: 删除指定任务及其关联的临时文件目录。
    """
    try:
        success = _delete_invoice_task_db(task_id)
        if not success:
            return jsonify({"status": "error", "message": "任务不存在"}), 404
        return jsonify({"status": "success", "message": "任务已删除"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize/tasks/<task_id>/cancel', methods=['POST'])
@login_required
@permission_required('amazon_invoice_export:export')
def cancel_invoice_organize_task(task_id):
    """
    取消发票整理任务

    简介: 将 pending 状态的任务标记为 cancelled，已开始执行的任务不可取消。
    """
    try:
        success, message = _cancel_invoice_task_db(task_id)
        if not success:
            return jsonify({"status": "error", "message": message}), 400
        return jsonify({"status": "success", "message": "任务已取消"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@amazon_invoice_export_bp.route('/amazon/invoices/organize/tasks/<task_id>/retry', methods=['POST'])
@login_required
@permission_required('amazon_invoice_export:export')
def retry_invoice_organize_task(task_id):
    """
    重试发票整理任务

    简介: 基于原任务参数创建新任务并提交，仅 failed 或 cancelled 状态可重试。

    返回:
        { status, message, data: { task_id } }   task_id 为新任务 ID
    """
    try:
        success, result = _retry_invoice_task_db(task_id)
        if not success:
            return jsonify({"status": "error", "message": result}), 400
        return jsonify({"status": "success", "message": "任务已重新提交", "data": {"task_id": result}})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
# 辅助函数 — 单位转换
# ============================================================

def _convert_length(value, unit):
    """
    将箱规长度统一转换为厘米

    简介: 支持英寸 (IN) 和厘米 (CM) 两种单位。

    详细:
      - IN: value * 2.54 → cm
      - CM: 原值返回
      - 其他单位或空值: 原值返回
    """
    if value is None or unit is None:
        return None
    unit = unit.upper()
    if unit == "IN":
        return float(value) * 2.54
    if unit == "CM":
        return float(value)
    return float(value)


def _convert_weight(value, unit):
    """
    将箱规重量统一转换为千克

    简介: 支持磅 (LB) 和千克 (KG) 两种单位及其常见别名。

    详细:
      - LB / LBS / POUND / POUNDS: value * 0.453592 → kg
      - KG / KGS / KILOGRAM / KILOGRAMS: 原值返回
      - 其他单位或空值: 原值返回
    """
    if value is None or unit is None:
        return None
    unit = unit.upper()
    if unit in ("LB", "LBS", "POUND", "POUNDS"):
        return float(value) * 0.453592
    if unit in ("KG", "KGS", "KILOGRAM", "KILOGRAMS"):
        return float(value)
    return float(value)


# ============================================================
# 辅助函数 — 参数校验
# ============================================================

def _require_shop_id() -> int:
    """
    从请求参数中提取并校验 shop_id

    简介: 校验 shop_id 存在且为整数，否则抛出 ValueError。

    详细:
      - 从 GET 参数 `shop_id` 中获取
      - 为空或非整数时抛出带有中文提示的 ValueError
    """
    shop_id = request.args.get('shop_id', '').strip() or None
    if not shop_id:
        raise ValueError("缺少必要参数: shop_id")
    try:
        return int(shop_id)
    except ValueError:
        raise ValueError("shop_id 必须是整数")


# ============================================================
# 辅助函数 — 数据查询
# ============================================================

def _fetch_shipment_invoice_data(shop_id, shipment_id):
    """
    查询货件发票导出数据

    简介: 根据店铺和货件 ID 查询箱规、关联产品信息，组装为模板填充用的数据结构。

    详细:
      1. 从 amazon_inbound_shipments_detail 获取仓库编号和 Amazon 参考号
      2. 从 amazon_inbound_plan_boxes 获取所有箱规（含 items_json）
      3. 收集所有 msku，批量查询 products 表获取产品详情
      4. 逐箱逐 item 计算总数量和总货值，整合返回

    返回:
        (warehouse_id, [dict, ...])
        warehouse_id: 目标仓库编号
        list: 每箱每 SKU 一行的填充数据
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT destination_warehouse_id, amazon_reference_id FROM amazon_inbound_shipments_detail WHERE shop_id = %s AND shipment_confirmation_id = %s",
                (shop_id, shipment_id),
            )
            shipment_row = cursor.fetchone()
            warehouse_id = shipment_row["destination_warehouse_id"] if shipment_row else ""
            amazon_reference_id = shipment_row["amazon_reference_id"] if shipment_row else ""

            cursor.execute(
                """SELECT * FROM amazon_inbound_plan_boxes
                   WHERE shop_id = %s AND shipment_id = %s
                   ORDER BY id ASC""",
                (shop_id, shipment_id),
            )
            boxes = cursor.fetchall()

            if not boxes:
                return warehouse_id, []

            msku_set = set()
            box_items = []
            for box in boxes:
                items = json.loads(box["items_json"] or "[]")
                for item in items:
                    msku = item.get("msku")
                    if msku:
                        msku_set.add(msku)
                    box_items.append((box, item))

            product_map = {}
            if msku_set:
                placeholders = ",".join(["%s"] * len(msku_set))
                cursor.execute(
                    f"""SELECT seller_sku, declare_name_en, declare_name_cn,
                               material_en, material_cn, purpose, brand, model,
                               declare_value, currency, hs_code, is_electric, is_magnetic,
                               amazon_internal_id, vat_number, eori_number, sales_url,
                               weight_kg, dimensions_cm, asin, image_url, fnsku
                        FROM products WHERE seller_sku IN ({placeholders})""",
                    tuple(msku_set),
                )
                for prod in cursor.fetchall():
                    product_map[prod["seller_sku"]] = prod

            results = []
            for box, item in box_items:
                msku = item.get("msku")
                prod = product_map.get(msku, {})
                box_qty = box.get("quantity") or 1
                item_qty = item.get("quantity") or 0
                total_qty = box_qty * item_qty
                unit_price = prod.get("declare_value") or Decimal("0")
                total_value = float(unit_price) * total_qty

                results.append({
                    "image_url": prod.get("image_url") or "",
                    "is_declare": "否",
                    "carton_count": box_qty,
                    "declare_name_en": prod.get("declare_name_en") or "",
                    "declare_name_cn": prod.get("declare_name_cn") or "",
                    "material_en": prod.get("material_en") or "",
                    "material_cn": prod.get("material_cn") or "",
                    "purpose": prod.get("purpose") or "",
                    "brand": prod.get("brand") or "",
                    "model": prod.get("model") or "",
                    "unit_quantity": item_qty,
                    "total_quantity": total_qty,
                    "unit_declare_value": float(unit_price) if unit_price else 0,
                    "total_declare_value": total_value,
                    "currency": prod.get("currency") or "USD",
                    "carton_weight_kg": (
                        round(_convert_weight(box.get("weight_value"), box.get("weight_unit")), 3)
                        if _convert_weight(box.get("weight_value"), box.get("weight_unit"))
                        else ""
                    ),
                    "hs_code": prod.get("hs_code") or "",
                    "is_electric": "是" if prod.get("is_electric") else "否",
                    "is_magnetic": "是" if prod.get("is_magnetic") else "否",
                    "fba_box_id": box.get("box_id") or "",
                    "amazon_reference_id": amazon_reference_id or "",
                    "carton_length_cm": (
                        round(_convert_length(box.get("dimensions_length"), box.get("dimensions_unit")), 2)
                        if _convert_length(box.get("dimensions_length"), box.get("dimensions_unit"))
                        else ""
                    ),
                    "carton_width_cm": (
                        round(_convert_length(box.get("dimensions_width"), box.get("dimensions_unit")), 2)
                        if _convert_length(box.get("dimensions_width"), box.get("dimensions_unit"))
                        else ""
                    ),
                    "carton_height_cm": (
                        round(_convert_length(box.get("dimensions_height"), box.get("dimensions_unit")), 2)
                        if _convert_length(box.get("dimensions_height"), box.get("dimensions_unit"))
                        else ""
                    ),
                    "warehouse_code": warehouse_id,
                    "delivery_address": box.get("destination_region_state") or "",
                    "country": "US",
                    "vat_number": prod.get("vat_number") or "",
                    "eori_number": prod.get("eori_number") or "",
                    "sales_url": prod.get("sales_url") or "",
                    "product_weight_kg": float(prod.get("weight_kg")) if prod.get("weight_kg") else "",
                    "product_dimensions": prod.get("dimensions_cm") or "",
                    "asin": prod.get("asin") or "",
                    "fnsku": prod.get("fnsku") or "",
                })

            return warehouse_id, results
    finally:
        conn.close()


def _get_invoice_config(provider_id):
    """
    获取货代发票模板配置

    简介: 从 logistics_invoice_configs 表查询货代对应的模板文件、表头行、数据起始行和字段映射。

    详细:
      - field_mappings 在库中为 JSON 字符串，读取时自动解析为 list
      - 拼接完整的 template_path（TEMPLATES_DIR + template_file）

    返回:
        dict 或 None（未找到配置时）:
            { provider_id, template_file, template_path, header_row, data_start_row, field_mappings }
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT provider_id, template_file, header_row, data_start_row, field_mappings
                   FROM logistics_invoice_configs WHERE provider_id = %s""",
                (provider_id,)
            )
            row = cursor.fetchone()
            if not row:
                return None
            row['field_mappings'] = json.loads(row['field_mappings']) if isinstance(row['field_mappings'], str) else row['field_mappings']
            row['template_path'] = os.path.join(TEMPLATES_DIR, row['template_file'])
            return row
    finally:
        conn.close()


def _get_shipments_by_inbound_plan(shop_id, inbound_plan_id):
    """
    查询入库计划下的所有货件

    简介: 用于批量整理任务，获取指定入库计划中每个货件的 shipment_id 和 shipment_confirmation_id。

    返回:
        list of dict: [{ shipment_id, shipment_confirmation_id }, ...]
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT s.shipment_id, d.shipment_confirmation_id
                   FROM amazon_inbound_shipments s
                   LEFT JOIN amazon_inbound_shipments_detail d
                       ON s.shipment_id = d.shipment_id AND d.shop_id = s.shop_id
                   WHERE s.shop_id = %s AND s.inbound_plan_id = %s
                   ORDER BY s.shipment_id""",
                (shop_id, inbound_plan_id)
            )
            return cursor.fetchall()
    finally:
        conn.close()


# ============================================================
# 辅助函数 — Excel 构建
# ============================================================

def _build_excel(data, provider_id):
    """
    基于货代配置和模板构建发票 xlsx

    简介: 加载货代上传的 Excel 模板，按字段映射填充数据行，并插入产品图片。

    详细:
      1. 读取 logistics_invoice_configs 中的模板路径和字段映射
      2. 解析模板表头，建立 header_name ↔ field_key 映射
      3. 逐行填充数据，应用边框和对齐样式
      4. 如果模板包含 image_url 列，下载/加载图片并插入对应单元格
         - 支持本地路径（/static/ 开头）和远程 URL
         - 图片自动缩放至 60x50px 缩略图，居中放置
      5. 返回 BytesIO 流

    返回:
        io.BytesIO: Excel 文件内存流
    """
    config = _get_invoice_config(provider_id)
    if not config:
        raise ValueError(f"未找到货代 {provider_id} 的发票配置")

    template_path = config['template_path']
    header_row = config['header_row']
    data_start_row = config['data_start_row']
    field_mappings = config['field_mappings']

    field_to_header = {m['field_key']: m['header_name'] for m in field_mappings}
    header_to_field = {m['header_name']: m['field_key'] for m in field_mappings}

    wb = load_workbook(template_path)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value:
            headers.append(str(cell_value))
        else:
            break

    img_col_idx = None
    for idx, h in enumerate(headers, 1):
        fk = header_to_field.get(h, '')
        if fk == 'image_url':
            img_col_idx = idx
            break

    if not headers:
        headers = sorted(field_mappings, key=lambda m: m.get('sort_order', 0))
        headers = [m['header_name'] for m in headers]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_offset, row_data in enumerate(data):
        row_idx = data_start_row + row_offset
        ws.row_dimensions[row_idx].height = 60
        for col_idx, header_name in enumerate(headers, 1):
            field_key = header_to_field.get(header_name)
            if field_key:
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field_key))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if img_col_idx:
        for row_offset, row_data in enumerate(data):
            row_idx = data_start_row + row_offset
            img_url = row_data.get("image_url")
            if not img_url:
                continue
            try:
                pil_img = None
                parsed = urlparse(img_url)
                path = parsed.path
                if path.startswith("/static/") or path.startswith("static/"):
                    local_path = path.lstrip("/")
                    pil_img = PILImage.open(local_path)
                else:
                    req = urllib.request.Request(
                        img_url, headers={"User-Agent": "Mozilla/5.0"}
                    )
                    with urllib.request.urlopen(req, timeout=15) as response:
                        image_bytes = response.read()
                    pil_img = PILImage.open(io.BytesIO(image_bytes))

                if pil_img.mode in ("RGBA", "LA", "P"):
                    background = PILImage.new("RGB", pil_img.size, (255, 255, 255))
                    if pil_img.mode == "P":
                        pil_img = pil_img.convert("RGBA")
                    if pil_img.mode in ("RGBA", "LA"):
                        background.paste(pil_img, mask=pil_img.split()[-1])
                        pil_img = background
                    else:
                        pil_img = pil_img.convert("RGB")

                max_width = 60
                max_height = 50
                pil_img.thumbnail((max_width, max_height), PILImage.LANCZOS)

                img_buf = io.BytesIO()
                pil_img.save(img_buf, format="PNG")
                img_buf.seek(0)
                xl_img = XLImage(img_buf)

                cell_width_px = 84
                cell_height_px = 80
                offset_x = max(0, (cell_width_px - xl_img.width) // 2)
                offset_y = max(0, (cell_height_px - xl_img.height) // 2)

                marker = AnchorMarker(
                    col=img_col_idx - 1,
                    row=row_idx - 1,
                    colOff=pixels_to_EMU(offset_x),
                    rowOff=pixels_to_EMU(offset_y),
                )
                size = XDRPositiveSize2D(
                    cx=pixels_to_EMU(xl_img.width),
                    cy=pixels_to_EMU(xl_img.height),
                )
                xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws._images.append(xl_img)

                cell = ws.cell(row=row_idx, column=img_col_idx)
                cell.value = None
            except Exception as e:
                print(f"[Invoice Export] 图片插入失败 [{img_url}]: {e}")
                pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# 辅助函数 — 任务 CRUD
# ============================================================

def _create_invoice_task(task_id, shop_id, inbound_plan_id, provider_id):
    """
    创建发票整理任务记录

    简介: 向 fba_invoice_tasks 表插入一条 pending 状态的任务。

    详细:
      - 任务有效期 24 小时（expired_at = now + 24h）
      - 初始状态 pending，进度 0
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            expired_at = datetime.now() + timedelta(hours=24)
            cursor.execute(
                """INSERT INTO fba_invoice_tasks (id, shop_id, inbound_plan_id, provider_id, status, progress, expired_at)
                   VALUES (%s, %s, %s, %s, 'pending', 0, %s)""",
                (task_id, shop_id, inbound_plan_id, provider_id, expired_at)
            )
            conn.commit()
    finally:
        conn.close()


def _update_invoice_task(task_id, status=None, progress=None, message=None, result_path=None, completed_at=None):
    """
    更新发票整理任务状态

    简介: 根据传入的非 None 参数动态构建 UPDATE 语句，更新指定的任务字段。

    详细:
      - 所有参数均为可选，仅更新传入的非 None 字段
      - 常用于进度上报、状态变更、结果路径记录等场景
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            fields = []
            values = []
            if status is not None:
                fields.append("status = %s")
                values.append(status)
            if progress is not None:
                fields.append("progress = %s")
                values.append(progress)
            if message is not None:
                fields.append("message = %s")
                values.append(message)
            if result_path is not None:
                fields.append("result_path = %s")
                values.append(result_path)
            if completed_at is not None:
                fields.append("completed_at = %s")
                values.append(completed_at)
            if not fields:
                return
            cursor.execute(
                f"UPDATE fba_invoice_tasks SET {', '.join(fields)} WHERE id = %s",
                tuple(values + [task_id])
            )
            conn.commit()
    finally:
        conn.close()


def _get_invoice_task(task_id):
    """
    根据 ID 查询单个发票整理任务

    返回:
        dict 或 None（不存在时）
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT id, shop_id, inbound_plan_id, provider_id, status, progress, message, result_path,
                          created_at, completed_at, expired_at
                   FROM fba_invoice_tasks WHERE id = %s""",
                (task_id,)
            )
            return cursor.fetchone()
    finally:
        conn.close()


def _list_invoice_tasks(shop_id, status=None, page=1, page_size=10):
    """
    分页查询发票整理任务列表

    简介: 按 shop_id 查询，支持 status 过滤和分页。

    返回:
        { list: [dict], total: int, page: int, page_size: int }
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            conditions = ["shop_id = %s"]
            params = [shop_id]
            if status:
                conditions.append("status = %s")
                params.append(status)
            where_clause = " AND ".join(conditions)

            cursor.execute(
                f"SELECT COUNT(*) as total FROM fba_invoice_tasks WHERE {where_clause}",
                tuple(params)
            )
            total = cursor.fetchone()['total']

            offset = (page - 1) * page_size
            cursor.execute(
                f"""SELECT id, shop_id, inbound_plan_id, provider_id, status, progress, message, result_path,
                           created_at, completed_at, expired_at
                    FROM fba_invoice_tasks
                    WHERE {where_clause}
                    ORDER BY created_at DESC
                    LIMIT %s OFFSET %s""",
                tuple(params + [page_size, offset])
            )
            return {"list": cursor.fetchall(), "total": total, "page": page, "page_size": page_size}
    finally:
        conn.close()


def _delete_invoice_task_db(task_id):
    """
    删除任务及关联文件

    简介: 先删除任务对应的临时文件目录，再删除数据库记录。

    详细:
      - 任务目录: static/shipping_invoice_templates/task/<task_id>/
      - 使用 shutil.rmtree 递归删除

    返回:
        bool: True 表示成功删除，False 表示任务不存在
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT result_path FROM fba_invoice_tasks WHERE id = %s", (task_id,))
            row = cursor.fetchone()
            if row:
                task_dir = os.path.join(INVOICE_TASK_DIR, task_id)
                if os.path.exists(task_dir):
                    shutil.rmtree(task_dir, ignore_errors=True)
                cursor.execute("DELETE FROM fba_invoice_tasks WHERE id = %s", (task_id,))
                conn.commit()
                return True
            return False
    finally:
        conn.close()


def _cancel_invoice_task_db(task_id):
    """
    取消 pending 状态的任务

    简介: 仅允许取消尚未开始执行的任务。

    返回:
        (bool, str): (是否成功, 错误消息或 None)
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT status FROM fba_invoice_tasks WHERE id = %s", (task_id,))
            row = cursor.fetchone()
            if not row:
                return False, "任务不存在"
            if row['status'] != 'pending':
                return False, "只能取消 pending 状态的任务"
            cursor.execute("UPDATE fba_invoice_tasks SET status = 'cancelled' WHERE id = %s", (task_id,))
            conn.commit()
            return True, None
    finally:
        conn.close()


def _retry_invoice_task_db(task_id):
    """
    重试失败或已取消的任务

    简介: 基于原任务参数创建新任务并提交，返回新任务 ID。

    详细:
      - 仅 failed 或 cancelled 状态可重试
      - 会先清理过期任务
      - 新任务 ID 为新的 UUID

    返回:
        (bool, str): (是否成功, 错误消息 或 新 task_id)
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT status FROM fba_invoice_tasks WHERE id = %s", (task_id,))
            row = cursor.fetchone()
            if not row:
                return False, "任务不存在"
            if row['status'] not in ('failed', 'cancelled'):
                return False, "只能重试 failed 或 cancelled 状态的任务"

            cursor.execute(
                """SELECT shop_id, inbound_plan_id, provider_id FROM fba_invoice_tasks WHERE id = %s""",
                (task_id,)
            )
            orig = cursor.fetchone()
            new_task_id = str(uuid.uuid4())
            _cleanup_old_invoice_tasks()
            _create_invoice_task(new_task_id, orig['shop_id'], orig['inbound_plan_id'], orig['provider_id'])
            conn.commit()
            return True, new_task_id
    finally:
        conn.close()


# ============================================================
# 辅助函数 — 后台 Worker
# ============================================================

def _pick_pending_invoice_task():
    """
    取出一条 pending 任务并标记为 running

    简介: 按创建时间升序取最早的一条 pending 任务，原子性标记为 running。

    详细:
      - 使用 SELECT + UPDATE 实现简单的任务分发
      - 返回的任务信息供 Worker 执行使用

    返回:
        dict 或 None（无待处理任务时）
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT id, shop_id, inbound_plan_id, provider_id, status, progress, message, result_path,
                          created_at, completed_at, expired_at
                   FROM fba_invoice_tasks
                   WHERE status = 'pending'
                   ORDER BY created_at ASC
                   LIMIT 1"""
            )
            row = cursor.fetchone()
            if row:
                cursor.execute("UPDATE fba_invoice_tasks SET status = 'running' WHERE id = %s", (row['id'],))
                conn.commit()
            return row
    finally:
        conn.close()


def _reset_running_invoice_tasks():
    """
    重置遗留的 running 任务为 pending

    简介: Worker 启动时调用，将上次异常中断的 running 任务恢复为 pending 重新执行。
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("UPDATE fba_invoice_tasks SET status = 'pending' WHERE status = 'running'")
            affected = cursor.rowcount
            conn.commit()
            if affected:
                print(f"[Invoice Worker] 已重置 {affected} 个 running 任务为 pending")
    finally:
        conn.close()


def _cleanup_old_invoice_tasks():
    """
    清理过期或超时的旧任务

    简介: 删除已过期或超过 24 小时的已完成/失败任务及其临时文件。

    详细:
      - 清理条件: expired_at < NOW() 或 (created_at < 24h前 且状态为 completed/failed)
      - 同时删除对应磁盘上的 task 目录
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT id, result_path FROM fba_invoice_tasks
                   WHERE expired_at < NOW()
                      OR (created_at < DATE_SUB(NOW(), INTERVAL 24 HOUR)
                          AND status IN ('completed', 'failed'))"""
            )
            old_tasks = cursor.fetchall()
            for t in old_tasks:
                task_dir = os.path.join(INVOICE_TASK_DIR, t['id'])
                if os.path.exists(task_dir):
                    shutil.rmtree(task_dir, ignore_errors=True)
            if old_tasks:
                ids = [t['id'] for t in old_tasks]
                placeholders = ','.join(['%s'] * len(ids))
                cursor.execute(f"DELETE FROM fba_invoice_tasks WHERE id IN ({placeholders})", ids)
                conn.commit()
    finally:
        conn.close()


def _execute_invoice_task(task_id, shop_id, inbound_plan_id, provider_id):
    """
    执行发票整理任务

    简介: 遍历入库计划下的所有货件，逐个生成 xlsx 并打包为 zip。

    详细:
      1. 标记任务为 running，进度 0
      2. 查询入库计划下所有货件
      3. 逐货件调用 _fetch_shipment_invoice_data + _build_excel
      4. 每个 xlsx 写入临时目录
      5. 全部完成后打包为 zip，标记 completed
      6. 异常时标记 failed 并记录错误信息
    """
    try:
        _update_invoice_task(task_id, status='running', progress=0)

        shipments = _get_shipments_by_inbound_plan(shop_id, inbound_plan_id)
        if not shipments:
            _update_invoice_task(task_id, status='failed', message='该入库计划下没有货件')
            return

        total = len(shipments)
        task_dir = os.path.join(INVOICE_TASK_DIR, task_id)
        os.makedirs(task_dir, exist_ok=True)
        zip_path = os.path.join(task_dir, f"{inbound_plan_id}_invoices.zip")

        import zipfile
        xlsx_files = []

        for idx, shipment in enumerate(shipments):
            shipment_id = shipment['shipment_confirmation_id'] or shipment['shipment_id']
            try:
                _, data = _fetch_shipment_invoice_data(shop_id=shop_id, shipment_id=shipment_id)
                if not data:
                    print(f"[Invoice Worker] 货件 {shipment_id} 没有箱子数据，跳过")
                    continue

                xlsx_io = _build_excel(data, provider_id=provider_id)
                xlsx_path = os.path.join(task_dir, f"{shipment_id}.xlsx")
                with open(xlsx_path, 'wb') as f:
                    f.write(xlsx_io.getvalue())
                xlsx_files.append(xlsx_path)

                progress = int((idx + 1) / total * 100)
                _update_invoice_task(task_id, progress=progress,
                                     message=f'已处理 {idx + 1}/{total} 个货件')
            except Exception as e:
                print(f"[Invoice Worker] 货件 {shipment_id} 处理异常: {e}")

        if not xlsx_files:
            _update_invoice_task(task_id, status='failed', message='没有可导出的数据')
            return

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for xlsx_path in xlsx_files:
                zf.write(xlsx_path, os.path.basename(xlsx_path))

        _update_invoice_task(task_id, status='completed', progress=100,
                             result_path=zip_path, completed_at=datetime.now(),
                             message=f'共导出 {len(xlsx_files)} 个货件发票')
        print(f"[Invoice Worker] 任务 {task_id} 完成: {zip_path}")

    except Exception as e:
        print(f"[Invoice Worker] 任务 {task_id} 执行异常: {e}")
        import traceback
        traceback.print_exc()
        _update_invoice_task(task_id, status='failed', message=str(e)[:500])


def _invoice_worker_loop():
    """
    Worker 主循环

    简介: 死循环轮询 pending 任务，拾取后调用 _execute_invoice_task 执行。

    详细:
      - 有任务时执行，无任务时 sleep 3 秒
      - 执行异常仅标记失败，不影响循环继续
    """
    print("[Invoice Worker] 工作线程已启动")
    while True:
        try:
            task = _pick_pending_invoice_task()
            if task:
                print(f"[Invoice Worker] 开始执行任务: {task['id']}")
                try:
                    _execute_invoice_task(task['id'], task['shop_id'], task['inbound_plan_id'], task['provider_id'])
                except Exception as e:
                    print(f"[Invoice Worker] 任务 {task['id']} 执行异常: {e}")
                    import traceback
                    traceback.print_exc()
                    _update_invoice_task(task['id'], status='failed', message=str(e)[:500])
            else:
                time.sleep(3)
        except Exception as e:
            print(f"[Invoice Worker] 轮询异常: {e}")
            time.sleep(5)


def _start_invoice_workers():
    """
    启动后台 Worker 线程

    简介: 由 app.py 在应用启动时调用，重置遗留任务并启动守护线程。

    详细:
      - 先调用 _reset_running_invoice_tasks 清理上次中断的 running 任务
      - 以 daemon=True 启动线程，应用退出时自动终止
    """
    _reset_running_invoice_tasks()
    t = threading.Thread(target=_invoice_worker_loop, daemon=True)
    t.start()
    print("[Invoice Worker] 后台工作线程已启动")
