"""
货代管理与货代运单管理模块

路由一览:
  货代:
    GET     /api/logistics-providers                     货代列表（分页+搜索）
    POST    /api/logistics-providers                     创建货代
    GET     /api/logistics-providers/<id>                 货代详情
    PUT     /api/logistics-providers/<id>                 编辑货代
    PUT     /api/logistics-providers/batch-status         批量修改货代状态
    DELETE  /api/logistics-providers/<id>                 删除货代
  货代运单:
    GET     /api/logistics-waybills                      运单列表（分页+搜索）
    POST    /api/logistics-waybills                      创建运单
    GET     /api/logistics-waybills/<id>                  运单详情
    PUT     /api/logistics-waybills/<id>                  编辑运单
    PUT     /api/logistics-waybills/batch-status          批量修改运单状态
    DELETE  /api/logistics-waybills/<id>                  删除运单
    POST    /api/logistics-waybills/import                Excel 导入运单
    GET     /api/logistics-waybills/available-shipments   FBA 货件下拉列表
"""
# ============================================================
# 导入
# ============================================================
import io
import json
from datetime import datetime, date

import openpyxl
from flask import Blueprint, request, jsonify
from services.mysql_service import get_db_connection
from blueprints.user_auth import login_required, permission_required

logistics_bp = Blueprint('logistics', __name__, url_prefix='/api')

# ============================================================
# 常量
# ============================================================
WAYBILL_STATUS_INITIAL = 0    # 初始状态（草稿）
WAYBILL_STATUS_COMPLETED = 5  # 最终状态（已完成）


# ============================================================
# 工具函数
# ============================================================
def _get_conn():
    return get_db_connection()


def _val_or_none(val, cast_type=None):
    """如果值为 None 或空字符串则返回 None，否则按类型转换"""
    if val is None or val == '':
        return None
    if cast_type is not None:
        return cast_type(val)
    return val


def _sync_waybill_expense(conn, waybill_no, total_cost_cny, new_status):
    """
    根据新状态同步运单的支出记录：
      - 已完成 → 不存在则创建，已存在则更新金额
      - 非已完成 → 存在则删除
    """
    if not waybill_no:
        return
    if new_status == WAYBILL_STATUS_COMPLETED:
        from blueprints.expenses import create_expense_for_source
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT id, amount FROM expenses WHERE source_type = 'logistics_waybill' AND source_no = %s LIMIT 1",
                    (waybill_no,)
                )
                existing = cursor.fetchone()
                amount = float(total_cost_cny or 0)
                if existing:
                    if float(existing['amount']) != amount:
                        cursor.execute(
                            "UPDATE expenses SET amount = %s WHERE id = %s",
                            (amount, existing['id'])
                        )
                        conn.commit()
                    return
            create_expense_for_source(
                conn, '物流/头程',
                amount, datetime.now().strftime('%Y-%m-%d'),
                f"运单 {waybill_no}", 'logistics_waybill', waybill_no, 'company'
            )
        except Exception as e:
            print(f"[Logistics] 同步支出记录异常: {e}")
    else:
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    "DELETE FROM expenses WHERE source_type = 'logistics_waybill' AND source_no = %s",
                    (waybill_no,)
                )
                conn.commit()
        except Exception as e:
            print(f"[Logistics] 删除支出记录异常: {e}")


# ============================================================
# 路由: 货代 列表 / 详情 / 新增 / 编辑 / 批量状态 / 删除
# ============================================================

@logistics_bp.route('/logistics-providers', methods=['GET'])
@login_required
@permission_required('logistics_providers:page')
def list_providers():
    """查询货代列表（支持分页、搜索）"""
    try:
        keyword = request.args.get('keyword', '').strip() or None
        status = request.args.get('status', '').strip() or None
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))

        if page < 1:
            page = 1
        if page_size < 1 or page_size > 500:
            page_size = 20

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                conditions = ["1=1"]
                params = []

                if keyword:
                    conditions.append("(name LIKE %s OR contact_person LIKE %s OR phone LIKE %s)")
                    like_val = f"%{keyword}%"
                    params.extend([like_val, like_val, like_val])

                if status is not None:
                    conditions.append("status = %s")
                    params.append(int(status))

                where_clause = " AND ".join(conditions)

                cursor.execute(f"SELECT COUNT(*) as total FROM logistics_providers WHERE {where_clause}", tuple(params))
                total = cursor.fetchone()['total']

                offset = (page - 1) * page_size
                cursor.execute(f"""
                    SELECT id, name, contact_person, phone, email, address, remark, status, created_at, updated_at
                    FROM logistics_providers
                    WHERE {where_clause}
                    ORDER BY id DESC
                    LIMIT %s OFFSET %s
                """, tuple(params + [page_size, offset]))
                rows = cursor.fetchall()

                return jsonify({
                    "status": "success",
                    "data": {"list": rows, "total": total, "page": page, "page_size": page_size}
                })
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 查询货代异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-providers/<int:provider_id>', methods=['GET'])
@login_required
@permission_required('logistics_providers:page')
def get_provider(provider_id):
    """查询单个货代详情"""
    try:
        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT id, name, contact_person, phone, email, address, remark, status, created_at, updated_at
                    FROM logistics_providers WHERE id = %s
                """, (provider_id,))
                row = cursor.fetchone()

                if not row:
                    return jsonify({"status": "error", "message": "货代不存在"}), 404

                return jsonify({"status": "success", "data": row})
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 查询货代详情异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-providers', methods=['POST'])
@login_required
@permission_required('logistics_providers:create')
def create_provider():
    """创建货代"""
    try:
        data = request.get_json() or {}
        name = data.get('name', '').strip()
        if not name:
            return jsonify({"status": "error", "message": "货代名称不能为空"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO logistics_providers (name, contact_person, phone, email, address, remark, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (
                    name,
                    data.get('contact_person', '').strip() or None,
                    data.get('phone', '').strip() or None,
                    data.get('email', '').strip() or None,
                    data.get('address', '').strip() or None,
                    data.get('remark', '').strip() or None,
                    data.get('status', 1)
                ))
                conn.commit()
                new_id = cursor.lastrowid

                return jsonify({"status": "success", "message": "创建成功", "data": {"id": new_id}})
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 创建货代异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-providers/<int:provider_id>', methods=['PUT'])
@login_required
@permission_required('logistics_providers:edit')
def update_provider(provider_id):
    """更新货代"""
    try:
        data = request.get_json() or {}
        name = data.get('name', '').strip()
        if not name:
            return jsonify({"status": "error", "message": "货代名称不能为空"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    UPDATE logistics_providers
                    SET name=%s, contact_person=%s, phone=%s, email=%s, address=%s, remark=%s, status=%s
                    WHERE id = %s
                """, (
                    name,
                    data.get('contact_person', '').strip() or None,
                    data.get('phone', '').strip() or None,
                    data.get('email', '').strip() or None,
                    data.get('address', '').strip() or None,
                    data.get('remark', '').strip() or None,
                    data.get('status', 1),
                    provider_id
                ))
                conn.commit()

                if cursor.rowcount == 0:
                    return jsonify({"status": "error", "message": "货代不存在或无需更新"}), 404

                return jsonify({"status": "success", "message": "更新成功"})
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 更新货代异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-providers/batch-status', methods=['PUT'])
@login_required
@permission_required('logistics_providers:edit')
def batch_update_provider_status():
    """批量修改货代状态"""
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        status = data.get('status')

        if not ids or not isinstance(ids, list):
            return jsonify({"status": "error", "message": "请提供要修改的ID列表"}), 400
        if status is None:
            return jsonify({"status": "error", "message": "请提供目标状态"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                placeholders = ', '.join(['%s'] * len(ids))
                sql = f"UPDATE logistics_providers SET status = %s WHERE id IN ({placeholders})"
                cursor.execute(sql, tuple([int(status)] + [int(i) for i in ids]))
                conn.commit()
                return jsonify({
                    "status": "success",
                    "message": f"成功更新 {cursor.rowcount} 条记录",
                    "data": {"affected": cursor.rowcount}
                })
        finally:
            conn.close()
    except Exception as e:
        print(f"[Logistics] 批量更新货代状态异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-providers/<int:provider_id>', methods=['DELETE'])
@login_required
@permission_required('logistics_providers:delete')
def delete_provider(provider_id):
    """删除货代（有关联运单时禁止删除）"""
    try:
        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) as cnt FROM logistics_waybills WHERE provider_id = %s", (provider_id,))
                cnt = cursor.fetchone()['cnt']
                if cnt > 0:
                    return jsonify({"status": "error", "message": f"该货代存在 {cnt} 个运单，无法删除"}), 400

                cursor.execute("DELETE FROM logistics_providers WHERE id = %s", (provider_id,))
                conn.commit()

                if cursor.rowcount == 0:
                    return jsonify({"status": "error", "message": "货代不存在"}), 404

                return jsonify({"status": "success", "message": "删除成功"})
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 删除货代异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
# 路由: 货代运单 列表 / 详情 / 新增 / 编辑 / 批量状态 / 删除
# ============================================================

@logistics_bp.route('/logistics-waybills', methods=['GET'])
@login_required
@permission_required('logistics_waybills:page')
def list_waybills():
    """查询运单列表（支持分页、搜索）"""
    try:
        provider_id = request.args.get('provider_id', '').strip() or None
        status = request.args.get('status', '').strip() or None
        keyword = request.args.get('keyword', '').strip() or None
        waybill_no = request.args.get('waybill_no', '').strip() or None
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))

        if page < 1:
            page = 1
        if page_size < 1 or page_size > 500:
            page_size = 20

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                conditions = ["1=1"]
                params = []

                if provider_id:
                    conditions.append("wb.provider_id = %s")
                    params.append(int(provider_id))

                if status is not None:
                    conditions.append("wb.status = %s")
                    params.append(int(status))

                if waybill_no:
                    conditions.append("wb.waybill_no = %s")
                    params.append(waybill_no)

                if keyword:
                    conditions.append(
                        "(wb.waybill_no LIKE %s OR wb.shipment_id LIKE %s OR p.name LIKE %s OR wb.route_name LIKE %s)"
                    )
                    like_val = f"%{keyword}%"
                    params.extend([like_val, like_val, like_val, like_val])

                where_clause = " AND ".join(conditions)

                cursor.execute(f"""
                    SELECT COUNT(*) as total FROM logistics_waybills wb
                    LEFT JOIN logistics_providers p ON wb.provider_id = p.id
                    WHERE {where_clause}
                """, tuple(params))
                total = cursor.fetchone()['total']

                offset = (page - 1) * page_size
                cursor.execute(f"""
                    SELECT wb.id, wb.waybill_no, wb.provider_id, p.name as provider_name,
                           wb.shipment_id, wb.transport_type, wb.route_name,
                           wb.departure_port, wb.destination_port, wb.destination_warehouse,
                           wb.total_weight_kg, wb.total_volume_cbm, wb.total_cartons,
                           wb.freight_cost_cny, wb.tax_cost_cny, wb.misc_cost_cny,
                           wb.total_cost_cny, wb.cost_per_kg, wb.currency, wb.status,
                           wb.ship_date, wb.eta_date, wb.arrival_date, wb.delivery_date,
                           wb.tracking_url, wb.remark, wb.created_at, wb.updated_at
                    FROM logistics_waybills wb
                    LEFT JOIN logistics_providers p ON wb.provider_id = p.id
                    WHERE {where_clause}
                    ORDER BY wb.id DESC
                    LIMIT %s OFFSET %s
                """, tuple(params + [page_size, offset]))
                rows = cursor.fetchall()

                # 批量附加货件明细（SKU + 中文名 + 数量）
                if rows:
                    shipment_ids = [row['shipment_id'] for row in rows if row['shipment_id']]
                    if shipment_ids:
                        placeholders2 = ', '.join(['%s'] * len(shipment_ids))
                        cursor.execute(f"""
                            SELECT shipment_id, items_json
                            FROM amazon_inbound_plan_boxes
                            WHERE shipment_id IN ({placeholders2})
                        """, tuple(shipment_ids))
                        boxes = cursor.fetchall()

                        # 按 shipment_id 聚合 SKU 数量
                        sku_qty_map = {}  # {shipment_id: {sku: qty}}
                        all_skus = set()
                        for box in boxes:
                            sid = box['shipment_id']
                            items = json.loads(box.get('items_json') or '[]')
                            if isinstance(items, list):
                                for it in items:
                                    sku = it.get('msku', '')
                                    qty = int(it.get('quantity', 0))
                                    if sku:
                                        sku_qty_map.setdefault(sid, {}).setdefault(sku, 0)
                                        sku_qty_map[sid][sku] += qty
                                        all_skus.add(sku)

                        # 批量查产品中文名
                        sku_name_map = {}
                        if all_skus:
                            sku_placeholders = ', '.join(['%s'] * len(all_skus))
                            cursor.execute(f"""
                                SELECT seller_sku, COALESCE(product_name, declare_name_cn, '') as product_name
                                FROM products WHERE seller_sku IN ({sku_placeholders})
                            """, tuple(all_skus))
                            for p in cursor.fetchall():
                                sku_name_map[p['seller_sku']] = p['product_name']

                        # 附加到行
                        for row in rows:
                            sid = row['shipment_id']
                            if sid in sku_qty_map:
                                row['items'] = [
                                    {
                                        'seller_sku': sku,
                                        'product_name': sku_name_map.get(sku, ''),
                                        'quantity': qty,
                                    }
                                    for sku, qty in sku_qty_map[sid].items()
                                ]
                            else:
                                row['items'] = []

                return jsonify({
                    "status": "success",
                    "data": {"list": rows, "total": total, "page": page, "page_size": page_size}
                })
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 查询运单异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-waybills/<int:waybill_id>', methods=['GET'])
@login_required
@permission_required('logistics_waybills:page')
def get_waybill(waybill_id):
    """查询单个运单详情"""
    try:
        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT wb.id, wb.waybill_no, wb.provider_id, p.name as provider_name,
                           wb.shipment_id, wb.transport_type, wb.route_name,
                           wb.departure_port, wb.destination_port, wb.destination_warehouse,
                           wb.total_weight_kg, wb.total_volume_cbm, wb.total_cartons,
                           wb.freight_cost_cny, wb.tax_cost_cny, wb.misc_cost_cny,
                           wb.total_cost_cny, wb.cost_per_kg, wb.currency, wb.status,
                           wb.ship_date, wb.eta_date, wb.arrival_date, wb.delivery_date,
                           wb.tracking_url, wb.remark, wb.created_at, wb.updated_at
                    FROM logistics_waybills wb
                    LEFT JOIN logistics_providers p ON wb.provider_id = p.id
                    WHERE wb.id = %s
                """, (waybill_id,))
                row = cursor.fetchone()

                if not row:
                    return jsonify({"status": "error", "message": "运单不存在"}), 404

                return jsonify({"status": "success", "data": row})
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 查询运单详情异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-waybills', methods=['POST'])
@login_required
@permission_required('logistics_waybills:create')
def create_waybill():
    """创建运单"""
    try:
        data = request.get_json() or {}
        waybill_no = data.get('waybill_no', '').strip() or None
        provider_id = data.get('provider_id')

        if provider_id is None:
            return jsonify({"status": "error", "message": "货代ID不能为空"}), 400

        shipment_id = data.get('shipment_id', '').strip()
        if not shipment_id:
            return jsonify({"status": "error", "message": "货件ID不能为空"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT id FROM logistics_providers WHERE id = %s", (provider_id,))
                if not cursor.fetchone():
                    return jsonify({"status": "error", "message": "货代不存在"}), 400

                fields = [
                    'waybill_no', 'provider_id', 'shipment_id', 'transport_type', 'route_name',
                    'departure_port', 'destination_port', 'destination_warehouse',
                    'total_weight_kg', 'total_volume_cbm', 'total_cartons',
                    'freight_cost_cny', 'tax_cost_cny', 'misc_cost_cny',
                    'total_cost_cny', 'cost_per_kg', 'currency', 'status',
                    'ship_date', 'eta_date', 'arrival_date', 'delivery_date',
                    'tracking_url', 'remark'
                ]
                values = [
                    waybill_no,
                    int(provider_id),
                    shipment_id,
                    data.get('transport_type', 1),
                    data.get('route_name', '').strip() or None,
                    data.get('departure_port', '').strip() or None,
                    data.get('destination_port', '').strip() or None,
                    data.get('destination_warehouse', '').strip() or None,
                    _val_or_none(data.get('total_weight_kg'), float),
                    _val_or_none(data.get('total_volume_cbm'), float),
                    _val_or_none(data.get('total_cartons'), int),
                    _val_or_none(data.get('freight_cost_cny'), float),
                    _val_or_none(data.get('tax_cost_cny'), float),
                    _val_or_none(data.get('misc_cost_cny'), float),
                    _val_or_none(data.get('total_cost_cny'), float),
                    _val_or_none(data.get('cost_per_kg'), float),
                    data.get('currency', '').strip() or None,
                    data.get('status', 0),
                    _parse_date(data.get('ship_date')),
                    _parse_date(data.get('eta_date')),
                    _parse_date(data.get('arrival_date')),
                    _parse_date(data.get('delivery_date')),
                    data.get('tracking_url', '').strip() or None,
                    data.get('remark', '').strip() or None,
                ]

                placeholders = ', '.join(['%s'] * len(fields))
                sql = f"INSERT INTO logistics_waybills ({', '.join(fields)}) VALUES ({placeholders})"
                cursor.execute(sql, tuple(values))
                conn.commit()
                new_id = cursor.lastrowid

                # 新建时已是最终状态，直接入账
                new_status = data.get('status', 0)
                if new_status == WAYBILL_STATUS_COMPLETED:
                    _sync_waybill_expense(
                        conn, waybill_no,
                        float(data.get('total_cost_cny') or 0),
                        WAYBILL_STATUS_COMPLETED
                    )

                return jsonify({"status": "success", "message": "创建成功", "data": {"id": new_id}})
        finally:
            conn.close()
    except Exception as e:
        print(f"[Logistics] 创建运单异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-waybills/<int:waybill_id>', methods=['PUT'])
@login_required
@permission_required('logistics_waybills:edit')
def update_waybill(waybill_id):
    """更新运单"""
    try:
        data = request.get_json() or {}
        waybill_no = data.get('waybill_no', '').strip() or None
        provider_id = data.get('provider_id')

        if provider_id is None:
            return jsonify({"status": "error", "message": "货代ID不能为空"}), 400

        shipment_id = data.get('shipment_id', '').strip()
        if not shipment_id:
            return jsonify({"status": "error", "message": "货件ID不能为空"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT id FROM logistics_providers WHERE id = %s", (provider_id,))
                if not cursor.fetchone():
                    return jsonify({"status": "error", "message": "货代不存在"}), 400

                # 获取当前运单信息（旧状态、旧单号，用于状态变更判断）
                cursor.execute("SELECT status, waybill_no FROM logistics_waybills WHERE id = %s", (waybill_id,))
                old_row = cursor.fetchone()
                if not old_row:
                    return jsonify({"status": "error", "message": "运单不存在"}), 404
                old_status = old_row['status']
                old_waybill_no = old_row['waybill_no']

                fields = [
                    'waybill_no', 'provider_id', 'shipment_id', 'transport_type', 'route_name',
                    'departure_port', 'destination_port', 'destination_warehouse',
                    'total_weight_kg', 'total_volume_cbm', 'total_cartons',
                    'freight_cost_cny', 'tax_cost_cny', 'misc_cost_cny',
                    'total_cost_cny', 'cost_per_kg', 'currency', 'status',
                    'ship_date', 'eta_date', 'arrival_date', 'delivery_date',
                    'tracking_url', 'remark'
                ]
                values = [
                    waybill_no,
                    int(provider_id),
                    shipment_id,
                    data.get('transport_type', 1),
                    data.get('route_name', '').strip() or None,
                    data.get('departure_port', '').strip() or None,
                    data.get('destination_port', '').strip() or None,
                    data.get('destination_warehouse', '').strip() or None,
                    _val_or_none(data.get('total_weight_kg'), float),
                    _val_or_none(data.get('total_volume_cbm'), float),
                    _val_or_none(data.get('total_cartons'), int),
                    _val_or_none(data.get('freight_cost_cny'), float),
                    _val_or_none(data.get('tax_cost_cny'), float),
                    _val_or_none(data.get('misc_cost_cny'), float),
                    _val_or_none(data.get('total_cost_cny'), float),
                    _val_or_none(data.get('cost_per_kg'), float),
                    data.get('currency', '').strip() or None,
                    data.get('status', 0),
                    _parse_date(data.get('ship_date')),
                    _parse_date(data.get('eta_date')),
                    _parse_date(data.get('arrival_date')),
                    _parse_date(data.get('delivery_date')),
                    data.get('tracking_url', '').strip() or None,
                    data.get('remark', '').strip() or None,
                ]

                set_clause = ', '.join([f"{f} = %s" for f in fields])
                sql = f"UPDATE logistics_waybills SET {set_clause} WHERE id = %s"
                cursor.execute(sql, tuple(values + [waybill_id]))
                conn.commit()

                new_status = data.get('status', 0)

                # 状态变更时同步支出记录
                if old_status != new_status:
                    wb_no = old_waybill_no or waybill_no
                    if not wb_no:
                        cursor.execute("SELECT waybill_no, total_cost_cny FROM logistics_waybills WHERE id = %s", (waybill_id,))
                        row = cursor.fetchone()
                        if row:
                            wb_no = row['waybill_no']
                            total_cost_cny = row['total_cost_cny']
                        else:
                            total_cost_cny = 0
                    else:
                        cursor.execute("SELECT total_cost_cny FROM logistics_waybills WHERE id = %s", (waybill_id,))
                        cost_row = cursor.fetchone()
                        total_cost_cny = float(cost_row['total_cost_cny'] or 0) if cost_row else 0
                    _sync_waybill_expense(conn, wb_no, total_cost_cny, new_status)

                return jsonify({"status": "success", "message": "更新成功"})
        finally:
            conn.close()
    except Exception as e:
        print(f"[Logistics] 更新运单异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-waybills/batch-status', methods=['PUT'])
@login_required
@permission_required('logistics_waybills:edit')
def batch_update_waybill_status():
    """批量修改运单状态"""
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        status = data.get('status')

        if not ids or not isinstance(ids, list):
            return jsonify({"status": "error", "message": "请提供要修改的ID列表"}), 400
        if status is None:
            return jsonify({"status": "error", "message": "请提供目标状态"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                placeholders = ', '.join(['%s'] * len(ids))
                int_ids = [int(i) for i in ids]
                new_status = int(status)

                sql = f"UPDATE logistics_waybills SET status = %s WHERE id IN ({placeholders})"
                cursor.execute(sql, tuple([new_status] + int_ids))
                conn.commit()

                # 状态变更时同步支出记录
                cursor.execute(f"""
                    SELECT waybill_no, total_cost_cny FROM logistics_waybills WHERE id IN ({placeholders})
                """, tuple(int_ids))
                for row in cursor.fetchall():
                    _sync_waybill_expense(conn, row['waybill_no'], row['total_cost_cny'], new_status)

                return jsonify({
                    "status": "success",
                    "message": f"成功更新 {cursor.rowcount} 条记录",
                    "data": {"affected": cursor.rowcount}
                })
        finally:
            conn.close()
    except Exception as e:
        print(f"[Logistics] 批量更新运单状态异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@logistics_bp.route('/logistics-waybills/<int:waybill_id>', methods=['DELETE'])
@login_required
@permission_required('logistics_waybills:delete')
def delete_waybill(waybill_id):
    """删除运单"""
    try:
        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("DELETE FROM logistics_waybills WHERE id = %s", (waybill_id,))
                conn.commit()

                if cursor.rowcount == 0:
                    return jsonify({"status": "error", "message": "运单不存在"}), 404

                return jsonify({"status": "success", "message": "删除成功"})
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 删除运单异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
# 工具: Excel 导入
# ============================================================

def _find_header_row(ws):
    """在 Excel 工作表中查找表头行号"""
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        if any(v and ('运单号' in str(v) or '货代' in str(v)) for v in row_values):
            return row_idx
    return 1


def _build_col_indices(headers):
    """根据表头建立字段名到列索引的映射"""
    index_map = {}
    for idx, header in enumerate(headers, 1):
        if header is None:
            continue
        h = str(header).strip()
        if not h:
            continue
        h_lower = h.lower()
        if '运单号' in h:
            index_map['waybill_no'] = idx
        elif '货代' in h:
            index_map['provider_name'] = idx
        elif '货件' in h or 'shipment' in h_lower:
            index_map['shipment_id'] = idx
        elif '运输' in h:
            index_map['transport_type'] = idx
        elif '路线' in h or 'route' in h_lower:
            index_map['route_name'] = idx
        elif '启运' in h or '起运' in h or 'departure' in h_lower:
            index_map['departure_port'] = idx
        elif '目的港' in h or 'destination_port' in h_lower:
            index_map['destination_port'] = idx
        elif '目的仓' in h or 'destination_warehouse' in h_lower or '仓库' in h:
            index_map['destination_warehouse'] = idx
        elif '重量' in h or 'weight' in h_lower:
            index_map['total_weight_kg'] = idx
        elif '体积' in h or 'volume' in h_lower or 'cbm' in h_lower:
            index_map['total_volume_cbm'] = idx
        elif '箱数' in h or '件数' in h or 'carton' in h_lower:
            index_map['total_cartons'] = idx
        elif '运费' in h or 'freight' in h_lower:
            index_map['freight_cost_cny'] = idx
        elif '税' in h or 'tax' in h_lower:
            index_map['tax_cost_cny'] = idx
        elif '杂费' in h or 'misc' in h_lower or '其他费' in h:
            index_map['misc_cost_cny'] = idx
        elif '总费' in h or '总成本' in h or '合计' in h:
            index_map['total_cost_cny'] = idx
        elif '单价' in h or 'cost_per_kg' in h_lower:
            index_map['cost_per_kg'] = idx
        elif '币种' in h or 'currency' in h_lower:
            index_map['currency'] = idx
        elif '状态' in h:
            index_map['status'] = idx
        elif '发货日' in h or 'ship_date' in h_lower or '出运日' in h:
            index_map['ship_date'] = idx
        elif '预计' in h or 'ETA' in h_lower or 'eta' in h_lower:
            index_map['eta_date'] = idx
        elif '到港' in h or 'arrival' in h_lower:
            index_map['arrival_date'] = idx
        elif '签收' in h or 'delivery' in h_lower or '派送' in h:
            index_map['delivery_date'] = idx
        elif '跟踪' in h or 'tracking' in h_lower:
            index_map['tracking_url'] = idx
        elif '备注' in h or 'remark' in h_lower:
            index_map['remark'] = idx
    return index_map


def _parse_date(val):
    """将各种日期值转为 'YYYY-MM-DD' 字符串，失败返回 None"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None

    date_formats = [
        '%a, %d %b %Y %H:%M:%S %Z',  # 'Mon, 18 May 2026 00:00:00 GMT'
        '%Y-%m-%dT%H:%M:%S',          # '2026-05-18T00:00:00'
        '%Y-%m-%d %H:%M:%S',          # '2026-05-18 00:00:00'
        '%Y/%m/%d',                    # '2026/05/18'
        '%Y-%m-%d',                    # '2026-05-18'
    ]
    from datetime import datetime as dt
    for fmt in date_formats:
        try:
            return dt.strptime(s.split('.')[0], fmt).strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            continue

    return s[:10]


def _parse_num(val, cast=float):
    """安全地将值转为数字，失败返回 None"""
    if val is None or val == '':
        return None
    try:
        return cast(val)
    except (ValueError, TypeError):
        return None


@logistics_bp.route('/logistics-waybills/import', methods=['POST'])
@login_required
@permission_required('logistics_waybills:import')
def import_waybills():
    """导入货代账单 Excel"""
    try:
        provider_id = request.form.get('provider_id', '').strip()
        if not provider_id:
            return jsonify({"status": "error", "message": "货代ID不能为空"}), 400

        provider_id = int(provider_id)
        file = request.files.get('file')
        if not file:
            return jsonify({"status": "error", "message": "请上传 Excel 文件"}), 400

        filename = file.filename.lower()
        if not filename.endswith(('.xlsx', '.xls')):
            return jsonify({"status": "error", "message": "仅支持 .xlsx / .xls 格式"}), 400

        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active

        header_row = _find_header_row(ws)
        headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
        index_map = _build_col_indices(headers)

        if not index_map:
            return jsonify({"status": "error", "message": "未识别到有效表头，请检查 Excel 格式"}), 400

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT id FROM logistics_providers WHERE id = %s", (provider_id,))
                if not cursor.fetchone():
                    return jsonify({"status": "error", "message": "货代不存在"}), 400

                import_count = 0
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    def get_cell(key):
                        col = index_map.get(key)
                        if col:
                            return ws.cell(row=row_idx, column=col).value
                        return None

                    waybill_no = str(get_cell('waybill_no') or '').strip() or None
                    shipment_id = str(get_cell('shipment_id') or '').strip() or None
                    if not waybill_no and not shipment_id:
                        continue

                    fields = [
                        'waybill_no', 'provider_id', 'shipment_id', 'transport_type', 'route_name',
                        'departure_port', 'destination_port', 'destination_warehouse',
                        'total_weight_kg', 'total_volume_cbm', 'total_cartons',
                        'freight_cost_cny', 'tax_cost_cny', 'misc_cost_cny',
                        'total_cost_cny', 'cost_per_kg', 'currency', 'status',
                        'ship_date', 'eta_date', 'arrival_date', 'delivery_date',
                        'tracking_url', 'remark'
                    ]
                    values = [
                        waybill_no,
                        provider_id,
                        shipment_id,
                        _parse_num(get_cell('transport_type'), int) or 1,
                        str(get_cell('route_name') or '').strip() or None,
                        str(get_cell('departure_port') or '').strip() or None,
                        str(get_cell('destination_port') or '').strip() or None,
                        str(get_cell('destination_warehouse') or '').strip() or None,
                        _parse_num(get_cell('total_weight_kg')),
                        _parse_num(get_cell('total_volume_cbm')),
                        _parse_num(get_cell('total_cartons'), int),
                        _parse_num(get_cell('freight_cost_cny')),
                        _parse_num(get_cell('tax_cost_cny')),
                        _parse_num(get_cell('misc_cost_cny')),
                        _parse_num(get_cell('total_cost_cny')),
                        _parse_num(get_cell('cost_per_kg')),
                        str(get_cell('currency') or '').strip() or None,
                        _parse_num(get_cell('status'), int) or 0,
                        _parse_date(get_cell('ship_date')),
                        _parse_date(get_cell('eta_date')),
                        _parse_date(get_cell('arrival_date')),
                        _parse_date(get_cell('delivery_date')),
                        str(get_cell('tracking_url') or '').strip() or None,
                        str(get_cell('remark') or '').strip() or None,
                    ]

                    placeholders = ', '.join(['%s'] * len(fields))
                    sql = f"INSERT INTO logistics_waybills ({', '.join(fields)}) VALUES ({placeholders})"
                    cursor.execute(sql, tuple(values))
                    import_count += 1

                conn.commit()
                return jsonify({
                    "status": "success",
                    "message": f"导入成功，共 {import_count} 条运单",
                    "data": {"count": import_count}
                })
        finally:
            conn.close()

    except Exception as e:
        print(f"[Logistics] 导入运单异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
# 路由: FBA 货件下拉列表
# ============================================================

@logistics_bp.route('/logistics-waybills/available-shipments', methods=['GET'])
@login_required
def list_available_shipments():
    """
    查询可供绑定的 FBA 货件列表
    查询参数（可选）:
        status_list       - 逗号分隔的状态，默认 WORKING,SHIPPED
        keyword           - 搜索 shipment_id / shipment_name / destination_fulfillment_center_id
        exclude_waybill_id- 可选，当前正在编辑的运单ID
    """
    try:
        status_list = request.args.get('status_list', 'WORKING,SHIPPED').strip()
        keyword = request.args.get('keyword', '').strip() or None
        exclude_waybill_id = request.args.get('exclude_waybill_id', '').strip() or None

        statuses = [s.strip() for s in status_list.split(',') if s.strip()]
        if not statuses:
            statuses = ['WORKING', 'SHIPPED']

        conn = _get_conn()
        try:
            with conn.cursor() as cursor:
                conditions = ["1=1"]
                params = []

                placeholders = ', '.join(['%s'] * len(statuses))
                conditions.append(f"status IN ({placeholders})")
                params.extend(statuses)

                if keyword:
                    conditions.append(
                        "(shipment_confirmation_id LIKE %s OR name LIKE %s OR destination_warehouse_id LIKE %s)"
                    )
                    like_val = f"%{keyword}%"
                    params.extend([like_val, like_val, like_val])

                exclude_sub = """
                    shipment_confirmation_id NOT IN (
                        SELECT shipment_id FROM logistics_waybills
                        WHERE shipment_id IS NOT NULL
                    )
                """
                if exclude_waybill_id:
                    exclude_sub = """
                        (
                            shipment_confirmation_id NOT IN (
                                SELECT shipment_id FROM logistics_waybills
                                WHERE shipment_id IS NOT NULL AND id != %s
                            )
                            OR shipment_confirmation_id = (SELECT shipment_id FROM logistics_waybills WHERE id = %s)
                        )
                    """
                    params.extend([exclude_waybill_id, exclude_waybill_id])

                conditions.append(exclude_sub)
                where_clause = " AND ".join(conditions)

                cursor.execute(f"""
                    SELECT shipment_confirmation_id as shipment_id, name as shipment_name, status as shipment_status, destination_warehouse_id as destination_fulfillment_center_id
                    FROM amazon_inbound_shipments_detail
                    WHERE {where_clause}
                    ORDER BY sync_time DESC
                    LIMIT 500
                """, tuple(params))
                rows = cursor.fetchall()

                return jsonify({
                    "status": "success",
                    "data": {"list": rows}
                })
        finally:
            conn.close()
    except Exception as e:
        print(f"[Logistics] 查询可选货件异常: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
